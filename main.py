import base64
import csv
import io
import json
import os
import re
import urllib.error
import urllib.parse
import urllib.request
from types import SimpleNamespace
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from dotenv import load_dotenv
from knowledge_loader import load_knowledge_base

try:
    from greennode_agentbase import GreenNodeAgentBaseApp, PingStatus, RequestContext
    AGENTBASE_IMPORT_ERROR = None
except Exception as exc:  # Local Python 3.9 fallback; Docker uses Python 3.11.
    GreenNodeAgentBaseApp = None
    PingStatus = None
    RequestContext = Any
    AGENTBASE_IMPORT_ERROR = exc

try:
    from openai import OpenAI
except Exception:  # pragma: no cover - optional dependency at runtime
    OpenAI = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency at runtime
    PdfReader = None

try:
    from docx import Document
except Exception:  # pragma: no cover - optional dependency at runtime
    Document = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency at runtime
    load_workbook = None


load_dotenv()


def env_flag(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y", "on"}


# GitHub Pages calls the runtime directly from the browser, so the default
# runtime must expose normal HTTP routes with CORS enabled. The GreenNode SDK
# app can still be enabled explicitly if needed for platform-native tests.
USE_GREENNODE_AGENTBASE_APP = env_flag("USE_GREENNODE_AGENTBASE_APP", False)
app = GreenNodeAgentBaseApp() if (USE_GREENNODE_AGENTBASE_APP and GreenNodeAgentBaseApp) else None
ROOT = Path(__file__).resolve().parent
KB_DIR = ROOT / "knowledge_base"
KNOWLEDGE_BASE = load_knowledge_base(KB_DIR)

DEFAULT_TABLEAU_VIEWS = {
    "insurance_performance_csv": "d3bc93f4-3036-449c-84f8-0a4875cbfc5b",
    "insurance_traffic_csv": "d580ce59-6cc4-4e74-93e0-79b75b2f9154",
    "promotion_summary_csv": "5f8e8ee3-2a76-417b-91ab-e2f4e680e286",
}


def parse_csv_text(text: str) -> List[Dict[str, str]]:
    if not text or not text.strip():
        return []
    reader = csv.DictReader(io.StringIO(text.strip()))
    return [{k: (v or "").strip() for k, v in row.items()} for row in reader]


def decode_uploaded_data_url(data_url: str) -> bytes:
    if not data_url:
        return b""
    if "," in data_url:
        data_url = data_url.split(",", 1)[1]
    return base64.b64decode(data_url)


def extract_text_from_xlsx(content: bytes, max_rows: int = 80, max_cols: int = 18) -> str:
    if not load_workbook:
        return "[Chưa cài openpyxl nên chưa đọc được file Excel.]"
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    chunks = []
    for sheet in workbook.worksheets[:4]:
        chunks.append(f"[Sheet: {sheet.title}]")
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_index > max_rows:
                chunks.append("...")
                break
            values = ["" if value is None else str(value) for value in row[:max_cols]]
            if any(value.strip() for value in values):
                chunks.append(",".join(values))
    return "\n".join(chunks).strip()


def extract_text_from_pdf(content: bytes, max_pages: int = 8) -> str:
    if not PdfReader:
        return "[Chưa cài pypdf nên chưa đọc được file PDF.]"
    reader = PdfReader(io.BytesIO(content))
    pages = []
    for page in reader.pages[:max_pages]:
        pages.append(page.extract_text() or "")
    return "\n".join(pages).strip()


def extract_text_from_docx(content: bytes) -> str:
    if not Document:
        return "[Chưa cài python-docx nên chưa đọc được file Word.]"
    document = Document(io.BytesIO(content))
    paragraphs = [paragraph.text for paragraph in document.paragraphs if paragraph.text.strip()]
    return "\n".join(paragraphs).strip()


def extract_uploaded_file_text(file_info: Dict[str, Any], max_chars: int = 14000) -> Dict[str, Any]:
    name = str(file_info.get("name") or "uploaded-file")
    lower_name = name.lower()
    file_type = str(file_info.get("type") or "")
    text = str(file_info.get("text") or "")
    error = file_info.get("error")
    if error:
        return {"name": name, "type": file_type, "available": False, "error": error}
    try:
        content = decode_uploaded_data_url(str(file_info.get("data_url") or ""))
        if not text:
            if lower_name.endswith((".csv", ".txt")) or file_type.startswith("text/"):
                text = content.decode("utf-8-sig", errors="replace")
            elif lower_name.endswith((".xlsx", ".xlsm")):
                text = extract_text_from_xlsx(content)
            elif lower_name.endswith(".pdf"):
                text = extract_text_from_pdf(content)
            elif lower_name.endswith(".docx"):
                text = extract_text_from_docx(content)
            elif lower_name.endswith(".doc"):
                text = "[File .doc đời cũ chưa được hỗ trợ đọc trực tiếp; vui lòng đổi sang .docx hoặc PDF.]"
            else:
                text = "[Định dạng file chưa được hỗ trợ đọc trực tiếp.]"
    except Exception as exc:
        return {"name": name, "type": file_type, "available": False, "error": safe_exception_message(exc, 300)}

    text = text.strip()
    return {
        "name": name,
        "type": file_type,
        "available": bool(text),
        "text": text[:max_chars],
        "truncated": len(text) > max_chars,
    }


def summarize_uploaded_files(files: Any) -> Dict[str, Any]:
    if not isinstance(files, list) or not files:
        return {"available": False, "files": []}
    extracted = [extract_uploaded_file_text(file_info) for file_info in files if isinstance(file_info, dict)]
    return {"available": any(item.get("available") for item in extracted), "files": extracted}


def env_value(*names: str, default: str = "") -> str:
    for name in names:
        value = os.getenv(name)
        if value:
            return value.strip().strip('"').strip("'")
    return default


def bool_env(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def normalize_llm_base_url(value: str) -> str:
    base_url = (value or "").strip().strip('"').strip("'").rstrip("/")
    for suffix in ("/chat/completions", "/messages"):
        if base_url.endswith(suffix):
            base_url = base_url[: -len(suffix)]
    return base_url


def normalize_llm_model(value: str) -> str:
    model = (value or "").strip().strip('"').strip("'")
    aliases = {
        "qwen 3.5 27b": "qwen/qwen3-5-27b",
        "qwen3.5-27b": "qwen/qwen3-5-27b",
        "qwen3-5-27b": "qwen/qwen3-5-27b",
        "qwen/qwen3.5-27b": "qwen/qwen3-5-27b",
    }
    return aliases.get(model.lower(), model)


def http_json(url: str, payload: Optional[Dict[str, Any]] = None, headers: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
    data = None
    request_headers = {"Accept": "application/json"}
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        request_headers["Content-Type"] = "application/json"
    if headers:
        request_headers.update(headers)
    req = urllib.request.Request(url, data=data, headers=request_headers, method="POST" if payload is not None else "GET")
    timeout_seconds = parse_number(os.getenv("TABLEAU_HTTP_TIMEOUT_SECONDS")) or 30
    with urllib.request.urlopen(req, timeout=timeout_seconds) as response:
        return json.loads(response.read().decode("utf-8"))


def http_text(url: str, headers: Optional[Dict[str, str]] = None) -> str:
    req = urllib.request.Request(url, headers=headers or {}, method="GET")
    timeout_seconds = parse_number(os.getenv("TABLEAU_HTTP_TIMEOUT_SECONDS")) or 30
    with urllib.request.urlopen(req, timeout=timeout_seconds) as response:
        return response.read().decode("utf-8-sig")


def safe_exception_message(exc: Exception, limit: int = 500) -> str:
    message = f"{exc.__class__.__name__}: {str(exc)}"
    if isinstance(exc, urllib.error.HTTPError):
        try:
            body = exc.read().decode("utf-8", errors="replace").strip()
            if body:
                body = re.sub(r"\s+", " ", body)
                message = f"{message}; body={body[:limit]}"
        except Exception:
            pass
    return message[:limit]


def tableau_credentials() -> Dict[str, str]:
    return {
        "server": env_value("TABLEAU_SERVER_URL", "SERVER").rstrip("/"),
        "site": env_value("TABLEAU_SITE_CONTENT_URL", "SITE_NAME"),
        "pat_name": env_value("TABLEAU_PAT_NAME", "PAT_NAME"),
        "pat_secret": env_value("TABLEAU_PAT_SECRET", "PAT_VALUE"),
        "api_version": env_value("TABLEAU_API_VERSION", "REST_API_VERSION", default="3.4"),
    }


def tableau_is_configured() -> bool:
    creds = tableau_credentials()
    return all(creds[key] for key in ["server", "site", "pat_name", "pat_secret", "api_version"])


def tableau_sign_in() -> Dict[str, str]:
    creds = tableau_credentials()
    sign_in_url = f"{creds['server']}/api/{creds['api_version']}/auth/signin"
    payload = {
        "credentials": {
            "personalAccessTokenName": creds["pat_name"],
            "personalAccessTokenSecret": creds["pat_secret"],
            "site": {"contentUrl": creds["site"]},
        }
    }
    result = http_json(sign_in_url, payload=payload)
    credentials = result["credentials"]
    return {
        "server": creds["server"],
        "api_version": creds["api_version"],
        "token": credentials["token"],
        "site_id": credentials["site"]["id"],
    }


def tableau_view_id(data_key: str) -> str:
    env_keys = {
        "insurance_performance_csv": "TABLEAU_VIEW_INSURANCE_PERFORMANCE_ID",
        "insurance_product_performance_csv": "TABLEAU_VIEW_PRODUCT_PERFORMANCE_ID",
        "insurance_traffic_csv": "TABLEAU_VIEW_INSURANCE_TRAFFIC_ID",
        "promotion_summary_csv": "TABLEAU_VIEW_PROMOTION_SUMMARY_ID",
    }
    return env_value(env_keys.get(data_key, ""), default=DEFAULT_TABLEAU_VIEWS.get(data_key, ""))


def app_id_from_message(message: str) -> Optional[str]:
    match = re.search(r"\b\d{4,}\b", message)
    return match.group(0) if match else None


def wants_zone_breakdown(message: str) -> bool:
    text = normalize_text(message)
    return "zone" in text or "zoneid" in text or "zone id" in text


def row_matches_app_id(row: Dict[str, str], app_id: str) -> bool:
    if not app_id:
        return False
    dimension_aliases: List[str] = []
    for key in ["app_id", "sku", "app_name", "product_name", "zone_id", "break_view"]:
        dimension_aliases.extend(DIMENSION_ALIASES[key])
    for column, value in row.items():
        if not value:
            continue
        if find_column({column: value}, dimension_aliases) is None:
            continue
        if app_id in digits_only(value) or app_id in normalize_text(value):
            return True
    return False


def rows_for_app_id(rows: List[Dict[str, str]], app_id: str) -> List[Dict[str, str]]:
    return [row for row in rows if row_matches_app_id(row, app_id)]


def requested_metric_keys(message: str) -> List[str]:
    text = normalize_text(message)
    requested = []
    if "tpv" in text or "doanh thu" in text or "total payment value" in text:
        requested.append("tpv")
    if "mpu" in text or "paying user" in text:
        requested.append("mpu")
    if "aov" in text or "average order value" in text:
        requested.append("aov")
    if not requested:
        requested = ["tpv", "mpu"]
    return requested


def metric_parts_for_answer(metrics: Dict[str, Optional[float]], requested_keys: List[str]) -> List[str]:
    labels = {
        "tpv": "TPV",
        "mpu": "MPU",
        "aov": "AOV",
    }
    parts = []
    for key in requested_keys:
        if key in labels:
            parts.append(f"{labels[key]} {format_number(metrics.get(key))}")
    return parts


def metric_change_parts(item: Dict[str, Any], requested_keys: List[str]) -> List[str]:
    parts = []
    if "tpv" in requested_keys:
        parts.append(f"TPV change {pct_text(item.get('tpv_change'))}")
    if "mpu" in requested_keys:
        parts.append(f"MPU change {pct_text(item.get('mpu_change'))}")
    return parts


def split_env_list(value: str) -> List[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def tableau_filter_attempts_for_message(message: str, data_key: str) -> List[Dict[str, str]]:
    app_id = app_id_from_message(message)
    if not app_id:
        return [{}]

    env_names = {
        "insurance_performance_csv": "TABLEAU_PERFORMANCE_APPID_FILTER_FIELDS",
        "insurance_product_performance_csv": "TABLEAU_PRODUCT_APPID_FILTER_FIELDS",
        "insurance_traffic_csv": "TABLEAU_TRAFFIC_APPID_FILTER_FIELDS",
        "promotion_summary_csv": "TABLEAU_PROMOTION_APPID_FILTER_FIELDS",
    }
    default_fields = {
        "insurance_performance_csv": "Sku,App ID",
        "insurance_product_performance_csv": "App ID,Sku,SKU Name",
        "insurance_traffic_csv": "SKU NAME,APP NAME,App ID",
        "promotion_summary_csv": "*App ID,App ID,Campaign Code,Campaign ID",
    }
    configured = env_value(
        env_names.get(data_key, ""),
        "TABLEAU_APPID_FILTER_FIELDS",
        "TABLEAU_APPID_FILTER_FIELD",
        "TABLEAU_SKU_FILTER_FIELD",
        default=default_fields.get(data_key, "Sku"),
    )
    attempts = [{field: app_id} for field in split_env_list(configured)]
    return attempts or [{}]


def fetch_tableau_view_csv(session: Dict[str, str], view_id: str, filters: Optional[Dict[str, str]] = None) -> str:
    query = {"maxAge": "1"}
    for field, value in (filters or {}).items():
        if value:
            query[f"vf_{field}"] = value
    query_string = urllib.parse.urlencode(query)
    url = (
        f"{session['server']}/api/{session['api_version']}/sites/{session['site_id']}"
        f"/views/{view_id}/data?{query_string}"
    )
    return http_text(url, headers={"X-Tableau-Auth": session["token"]})


def fetch_tableau_proxy_live_data(message: str) -> Dict[str, Any]:
    proxy_url = env_value("TABLEAU_PROXY_URL").rstrip("/")
    proxy_token = env_value("TABLEAU_PROXY_TOKEN")
    metadata = {"enabled": True, "used": False, "errors": [], "views": {}, "source": "proxy"}
    proxy_urls = [url.rstrip("/") for url in split_env_list(env_value("TABLEAU_PROXY_URLS"))]
    if proxy_url:
        proxy_urls.insert(0, proxy_url)
    proxy_urls = list(dict.fromkeys([url for url in proxy_urls if url]))
    if not proxy_urls:
        metadata["errors"].append("Thiếu TABLEAU_PROXY_URL.")
        return {"data": {}, "metadata": metadata}

    headers = {}
    if proxy_token:
        headers["Authorization"] = f"Bearer {proxy_token}"
    result: Dict[str, Any] = {}
    used_url = ""
    for candidate_url in proxy_urls:
        try:
            result = http_json(
                f"{candidate_url}/tableau/live",
                payload={"message": message},
                headers=headers,
            )
            used_url = candidate_url
            break
        except Exception as exc:
            metadata["errors"].append(
                f"Không gọi Tableau bridge {candidate_url} được: {safe_exception_message(exc, 500)}"
            )
    if not result:
        return {"data": {}, "metadata": metadata}

    data = result.get("data") or {}
    proxy_metadata = result.get("metadata") or {}
    metadata.update(proxy_metadata)
    metadata["source"] = "proxy"
    metadata["proxy_url_used"] = used_url
    metadata["used"] = bool(data)
    if not metadata["used"] and not metadata.get("errors"):
        metadata["errors"] = ["Tableau bridge không trả data."]
    return {"data": data, "metadata": metadata}


def fetch_tableau_live_data(message: str) -> Dict[str, Any]:
    if env_value("TABLEAU_PROXY_URL") and not bool_env("FORCE_DIRECT_TABLEAU", default=False):
        return fetch_tableau_proxy_live_data(message)

    metadata = {"enabled": True, "used": False, "errors": [], "views": {}}
    data: Dict[str, str] = {}
    if not tableau_is_configured():
        metadata["errors"].append("Thiếu Tableau env config nên chưa lấy live data.")
        return {"data": data, "metadata": metadata}

    try:
        session = tableau_sign_in()
    except Exception as exc:
        metadata["errors"].append(f"Không sign in Tableau được: {safe_exception_message(exc, 500)}")
        return {"data": data, "metadata": metadata}

    for data_key in [
        "insurance_performance_csv",
        "insurance_product_performance_csv",
        "insurance_traffic_csv",
        "promotion_summary_csv",
    ]:
        view_id = tableau_view_id(data_key)
        if not view_id:
            continue
        attempts = tableau_filter_attempts_for_message(message, data_key)
        empty_result: Optional[Dict[str, Any]] = None
        for filters in attempts:
            try:
                csv_text = fetch_tableau_view_csv(session, view_id, filters=filters)
                rows = parse_csv_text(csv_text)
                if filters and not rows and len(attempts) > 1:
                    empty_result = {"csv_text": csv_text, "filters": filters, "rows": rows}
                    metadata["errors"].append(f"{data_key} filters={filters}: Tableau returned 0 rows, trying next filter.")
                    continue
                filter_verified = True
                if filters and rows:
                    unfiltered_csv = fetch_tableau_view_csv(session, view_id, filters={})
                    if csv_text.strip() == unfiltered_csv.strip():
                        filter_verified = False
                        metadata["errors"].append(
                            f"{data_key} filters={filters}: filtered CSV equals unfiltered CSV, trying next filter."
                        )
                        if len(attempts) > 1:
                            continue
                data[data_key] = csv_text
                metadata["views"][data_key] = {
                    "view_id": view_id,
                    "rows": len(rows),
                    "filters": filters,
                    "filter_attempts": attempts,
                    "filter_verified": filter_verified,
                }
                metadata["used"] = True
                break
            except Exception as exc:
                metadata["errors"].append(
                    f"{data_key} filters={filters}: {safe_exception_message(exc, 500)}"
                )
        else:
            if empty_result and data_key not in data:
                data[data_key] = empty_result["csv_text"]
                metadata["views"][data_key] = {
                    "view_id": view_id,
                    "rows": 0,
                    "filters": empty_result["filters"],
                    "filter_attempts": attempts,
                }
    return {"data": data, "metadata": metadata}


def parse_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    raw = str(value).strip().replace(",", "").replace("%", "")
    if raw == "":
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def format_number(value: Optional[float], suffix: str = "") -> str:
    if value is None:
        return "N/A"
    if abs(value) >= 1_000_000_000:
        return f"{value / 1_000_000_000:.1f}B{suffix}"
    if abs(value) >= 1_000_000:
        return f"{value / 1_000_000:.1f}M{suffix}"
    if abs(value) >= 1_000:
        return f"{value:,.0f}{suffix}"
    return f"{value:.2f}{suffix}" if value % 1 else f"{value:,.0f}{suffix}"


def percent_change(current: Optional[float], previous: Optional[float]) -> Optional[float]:
    if current is None or previous in (None, 0):
        return None
    return (current - previous) / previous


def health_label(change: Optional[float]) -> str:
    if change is None:
        return "không đủ dữ liệu"
    if change >= 0:
        return "tốt"
    if change >= -0.05:
        return "tạm chấp nhận"
    if change <= -0.15:
        return "báo động"
    return "không ổn"


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def digits_only(value: Any) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def find_column(row: Dict[str, str], aliases: List[str]) -> Optional[str]:
    if not row:
        return None
    normalized_aliases = {normalize_text(alias) for alias in aliases}
    for column in row.keys():
        normalized_column = normalize_text(column)
        if normalized_column in normalized_aliases:
            return column
    for column in row.keys():
        normalized_column = normalize_text(column)
        if any(alias in normalized_column for alias in normalized_aliases):
            return column
    return None


def get_first_value(row: Dict[str, str], aliases: List[str]) -> str:
    column = find_column(row, aliases)
    return row.get(column, "") if column else ""


DIMENSION_ALIASES = {
    "app_id": ["App ID", "AppID", "APP ID", "*App ID", "app_id"],
    "sku": ["SKU", "Sku", "SKU Name", "Sku Name", "Switch Break View"],
    "app_name": ["App Name", "APP NAME", "App", "Application Name"],
    "product_name": ["Product Name", "PRODUCT NAME", "Product", "Category"],
    "campaign": ["Campaign Name", "Campaign", "Campaign View", "Campaign Code", "Campaign Source"],
    "source": ["FROM_SOURCE", "From Source", "Source", "Traffic Source"],
    "zone_id": ["ZONE ID", "Zone ID", "ZoneID"],
    "flow": ["FLOW", "Flow"],
    "break_view": ["Switch Break View", "BREAK BY DIMENSION", "Break View"],
}

METRIC_ALIASES = {
    "tpv": ["TPV", "Total Payment Value"],
    "mpu": ["MPU", "USERS_DAILY", "Users", "Paying Users", "Success Users"],
    "aov": ["AOV", "Average Order Value"],
    "transactions": ["TRANS", "Transaction", "Transactions", "Promo Trans"],
    "load": ["LOAD VALUE", "Access Users", "Load", "Traffic"],
    "cashier": ["Cashier Users", "Cashier", "Cashier Value"],
    "success": ["SUCCESS VALUE", "Success Users", "Success"],
    "success_rate": ["SUCCESS RATE", "%CR", "Conversion Rate"],
    "final_cost": ["*Final Cost", "Final Cost", "Cost"],
    "cost_tpv": ["%Cost/TPV", "Cost/TPV"],
    "promo_users": ["Promo Users"],
}

PERIOD_ALIASES = [
    "Period_Type",
    "Period",
    "Time",
    "Month",
    "Ymd",
    "Max. Ymd",
    "Max. report_date",
    "report_date",
    "Date",
]


def extract_metrics(row: Dict[str, str]) -> Dict[str, Optional[float]]:
    metrics = {}
    for metric, aliases in METRIC_ALIASES.items():
        metrics[metric] = parse_number(get_first_value(row, aliases))
    return {key: value for key, value in metrics.items() if value is not None}


def row_period(row: Dict[str, str]) -> str:
    return get_first_value(row, PERIOD_ALIASES) or "unknown"


def compact_row(row: Dict[str, str], max_fields: int = 14) -> Dict[str, str]:
    compact = {}
    for key, value in row.items():
        if value not in ("", None):
            compact[key] = value
        if len(compact) >= max_fields:
            break
    return compact


def table_profile(rows: List[Dict[str, str]]) -> Dict[str, Any]:
    if not rows:
        return {"available": False}
    return {
        "available": True,
        "row_count": len(rows),
        "columns": list(rows[0].keys()),
        "sample_rows": [compact_row(row) for row in rows[:8]],
    }


def build_breakdown(
    rows: List[Dict[str, str]],
    dimension_keys: List[str],
    limit: int = 20,
) -> Dict[str, Any]:
    if not rows:
        return {"available": False}

    first_row = rows[0]
    dimension_field = None
    dimension_key = None
    for key in dimension_keys:
        field = find_column(first_row, DIMENSION_ALIASES[key])
        if field:
            dimension_field = field
            dimension_key = key
            break

    if not dimension_field:
        return {
            "available": False,
            "reason": f"Không tìm thấy dimension trong các nhóm {', '.join(dimension_keys)}.",
            "columns": list(first_row.keys()),
        }

    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for index, row in enumerate(rows):
        group = row.get(dimension_field, "").strip() or "Unknown"
        grouped.setdefault(group, []).append(
            {
                "index": index,
                "period": row_period(row),
                "metrics": extract_metrics(row),
                "row": compact_row(row),
            }
        )

    items = []
    for group, points in grouped.items():
        points.sort(key=lambda item: item["index"])
        latest = points[-1]
        previous = points[-2] if len(points) >= 2 else None
        latest_metrics = latest["metrics"]
        previous_metrics = previous["metrics"] if previous else {}
        tpv_change = percent_change(latest_metrics.get("tpv"), previous_metrics.get("tpv"))
        mpu_change = percent_change(latest_metrics.get("mpu"), previous_metrics.get("mpu"))
        load_change = percent_change(latest_metrics.get("load"), previous_metrics.get("load"))
        success_change = percent_change(latest_metrics.get("success"), previous_metrics.get("success"))
        items.append(
            {
                "name": group,
                "latest_period": latest["period"],
                "previous_period": previous["period"] if previous else None,
                "latest": latest_metrics,
                "previous": previous_metrics,
                "tpv_change": tpv_change,
                "mpu_change": mpu_change,
                "load_change": load_change,
                "success_change": success_change,
                "health": health_label(tpv_change or success_change or load_change),
                "raw_latest": latest["row"],
            }
        )

    def ranking_value(item: Dict[str, Any]) -> float:
        latest = item.get("latest", {})
        return latest.get("tpv") or latest.get("mpu") or latest.get("success") or latest.get("load") or 0

    items.sort(key=ranking_value, reverse=True)
    return {
        "available": True,
        "dimension_key": dimension_key,
        "dimension_field": dimension_field,
        "items": items[:limit],
        "total_groups": len(items),
    }


def find_focus_matches(message: str, datasets: Dict[str, List[Dict[str, str]]]) -> List[Dict[str, Any]]:
    text = normalize_text(message)
    numeric_tokens = set(re.findall(r"\b\d{3,}\b", message))
    matches = []
    searchable_fields = []
    for aliases in DIMENSION_ALIASES.values():
        searchable_fields.extend(aliases)

    for dataset_name, rows in datasets.items():
        for row in rows:
            matched_fields = {}
            for column, value in row.items():
                if not value:
                    continue
                normalized_value = normalize_text(value)
                digit_value = digits_only(value)
                column_is_dimension = find_column({column: value}, searchable_fields) is not None
                text_match = len(normalized_value) >= 4 and normalized_value in text
                id_match = digit_value and digit_value in numeric_tokens
                if column_is_dimension and (text_match or id_match):
                    matched_fields[column] = value
            if matched_fields:
                matches.append(
                    {
                        "dataset": dataset_name,
                        "matched_fields": matched_fields,
                        "metrics": extract_metrics(row),
                        "period": row_period(row),
                        "row": compact_row(row),
                    }
                )
            if len(matches) >= 20:
                return matches
    return matches


def summarize_performance(rows: List[Dict[str, str]]) -> Dict[str, Any]:
    if not rows:
        return {"available": False}

    clean_rows = []
    for row in rows:
        clean_rows.append(
            {
                "period": row.get("Period_Type") or row.get("Period") or row.get("Time"),
                "break_view": row.get("Switch Break View") or row.get("BREAK BY DIMENSION"),
                "aov": parse_number(row.get("AOV")),
                "tpv": parse_number(row.get("TPV")),
                "transactions": parse_number(row.get("TRANS")),
                "users": parse_number(row.get("USERS_DAILY")),
                "cutoff": row.get("Max. Ymd") or row.get("Max. report_date"),
            }
        )

    latest = clean_rows[-1]
    previous = clean_rows[-2] if len(clean_rows) >= 2 else None
    tpv_change = percent_change(latest.get("tpv"), previous.get("tpv") if previous else None)
    user_change = percent_change(latest.get("users"), previous.get("users") if previous else None)

    return {
        "available": True,
        "latest": latest,
        "previous": previous,
        "tpv_change": tpv_change,
        "user_change": user_change,
        "health": health_label(tpv_change),
    }


def summarize_traffic(rows: List[Dict[str, str]]) -> Dict[str, Any]:
    if not rows:
        return {"available": False}

    crosstab_rows = normalize_traffic_crosstab_rows(rows)
    already_normalized_detail = any(
        row.get("Time") and (row.get("LOAD VALUE") or row.get("SUCCESS VALUE")) and (
            row.get("ZONE ID") or row.get("SKU NAME") or row.get("APP NAME") or row.get("FROM_SOURCE")
        )
        for row in rows
    )
    source_shape = "crosstab" if crosstab_rows or already_normalized_detail else "summary"
    if crosstab_rows:
        rows = crosstab_rows

    by_time: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        time_key = row.get("Time") or row.get("Max. report_date") or "unknown"
        current = by_time.setdefault(
            time_key,
            {
                "time": time_key,
                "load": None,
                "cashier": None,
                "success_rate": None,
                "success": None,
                "cutoff": row.get("Max. report_date"),
                "dimension": row.get("ZONE ID") or row.get("SKU NAME") or row.get("APP NAME") or row.get("FROM_SOURCE"),
            },
        )
        load = parse_number(row.get("LOAD VALUE"))
        cashier = parse_number(row.get("CASHIER VALUE") or row.get("Cashier Users"))
        success = parse_number(row.get("SUCCESS VALUE"))
        success_rate = parse_number(row.get("SUCCESS RATE"))
        current["load"] = load if load is not None else current.get("load")
        current["cashier"] = cashier if cashier is not None else current.get("cashier")
        current["success"] = success if success is not None else current.get("success")
        current["success_rate"] = success_rate if success_rate is not None else current.get("success_rate")
        current["cutoff"] = row.get("Max. report_date") or current.get("cutoff")

    for point in by_time.values():
        if point.get("success_rate") is None and point.get("load") not in (None, 0) and point.get("success") is not None:
            point["success_rate"] = point["success"] / point["load"]

    ordered = list(by_time.values())
    latest = ordered[-1]
    previous = ordered[-2] if len(ordered) >= 2 else None
    load_change = percent_change(latest.get("load"), previous.get("load") if previous else None)
    success_change = percent_change(latest.get("success"), previous.get("success") if previous else None)

    return {
        "available": True,
        "latest": latest,
        "previous": previous,
        "load_change": load_change,
        "success_change": success_change,
        "health": health_label(success_change),
        "source_shape": source_shape,
    }


def normalize_traffic_crosstab_rows(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    if not rows:
        return []
    first = rows[0]
    flow_col = find_column(first, DIMENSION_ALIASES["flow"])
    if not flow_col:
        return []
    date_columns = [
        column
        for column in first.keys()
        if re.search(r"\b\d{1,2}[-/][A-Za-z]{3}\b", column) or re.search(r"\b\d{1,2}[-/]\d{1,2}\b", column)
    ]
    if not date_columns:
        return []

    zone_col = find_column(first, DIMENSION_ALIASES["zone_id"])
    sku_col = find_column(first, DIMENSION_ALIASES["sku"])
    app_col = find_column(first, DIMENSION_ALIASES["app_name"])
    source_col = find_column(first, DIMENSION_ALIASES["source"])
    by_key: Dict[tuple, Dict[str, str]] = {}

    for row in rows:
        flow = normalize_text(row.get(flow_col))
        if "load" in flow:
            metric_col = "LOAD VALUE"
        elif "cashier" in flow:
            metric_col = "CASHIER VALUE"
        elif "success" in flow:
            metric_col = "SUCCESS VALUE"
        else:
            continue
        dimension = row.get(zone_col, "") if zone_col else ""
        sku = row.get(sku_col, "") if sku_col else ""
        app_name = row.get(app_col, "") if app_col else ""
        source = row.get(source_col, "") if source_col else ""
        for date_column in date_columns:
            key = (date_column, dimension, sku, app_name, source)
            record = by_key.setdefault(
                key,
                {
                    "Time": date_column,
                    "ZONE ID": dimension,
                    "SKU NAME": sku,
                    "APP NAME": app_name,
                    "FROM_SOURCE": source,
                },
            )
            record[metric_col] = row.get(date_column, "")
    return list(by_key.values())


def summarize_promotion(rows: List[Dict[str, str]]) -> Dict[str, Any]:
    if not rows:
        return {"available": False}

    grouped: Dict[str, Dict[str, float]] = {}
    for row in rows:
        group = row.get("C1") or row.get("C0") or "Unknown"
        measure = row.get("Measure Names")
        value = parse_number(row.get("Measure Values"))
        if not measure or value is None:
            continue
        grouped.setdefault(group, {})[measure] = value

    interesting = []
    for group, measures in grouped.items():
        tpv = measures.get("TPV")
        cost = measures.get("*Final Cost")
        cost_tpv = measures.get("%Cost/TPV")
        promo_users = measures.get("Promo Users")
        if tpv or cost or promo_users:
            interesting.append(
                {
                    "group": group,
                    "tpv": tpv,
                    "final_cost": cost,
                    "cost_tpv": cost_tpv,
                    "promo_users": promo_users,
                    "promo_trans": measures.get("Promo Trans"),
                    "cpu": measures.get("CPU"),
                    "cac": measures.get("CAC"),
                }
            )

    interesting.sort(key=lambda item: item.get("tpv") or 0, reverse=True)
    total = grouped.get("All") or grouped.get("ALL") or {}
    return {
        "available": True,
        "total": total,
        "top_groups": interesting[:8],
    }


def classify_intent(message: str, mode: str = "") -> str:
    normalized_mode = (mode or "").lower()
    text = message.lower()
    if normalized_mode in {"performance", "traffic"}:
        return normalized_mode
    if normalized_mode == "daily":
        return "performance"
    has_incentive_signal = any(
        keyword in text for keyword in ["promo", "promotion", "incentive", "cost", "campaign", "voucher", "chi phí", "khuyến mãi", "promotion code"]
    )
    if has_incentive_signal:
        return "incentive"
    if any(keyword in text for keyword in ["traffic", "funnel", "load", "cashier", "success", "access"]):
        return "traffic"
    if any(keyword in text for keyword in ["performance", "tpv", "mpu", "npu", "fpu", "aov", "doanh thu", "paying user"]):
        return "performance"
    has_product_signal = any(
        keyword in text for keyword in ["sản phẩm", "bao hiem", "bảo hiểm", "product", "sku", "appid", "app id", "chubb", "vbi", "tnds"]
    ) or bool(re.search(r"\b\d{4,}\b", message))
    if has_product_signal:
        return "performance"
    if normalized_mode in {"product", "traffic", "incentive", "report"}:
        return normalized_mode
    if any(keyword in text for keyword in ["report", "báo cáo", "monthly", "tháng"]):
        return "report"
    if any(keyword in text for keyword in ["cpu", "cac"]):
        return "incentive"
    return "performance"


def pct_text(change: Optional[float]) -> str:
    return format_number(change * 100, "%") if change is not None else "N/A"


def build_daily_answer(summaries: Dict[str, Any]) -> str:
    lines = ["Tóm tắt nhanh Insurance hôm nay"]
    perf = summaries.get("performance", {})
    traffic = summaries.get("traffic", {})
    promo = summaries.get("promotion", {})

    if perf.get("available"):
        latest = perf["latest"]
        previous = perf.get("previous") or {}
        lines.append(
            f"- Performance: {latest.get('period')} đạt TPV {format_number(latest.get('tpv'))}, "
            f"MPU {format_number(latest.get('users'))}; so với {previous.get('period', 'kỳ trước')} "
            f"TPV thay đổi {pct_text(perf.get('tpv_change'))}, trạng thái {perf.get('health')}."
        )

    if traffic.get("available"):
        latest = traffic["latest"]
        lines.append(
            f"- Traffic funnel: ngày {latest.get('time')} có load {format_number(latest.get('load'))}, "
            f"success {format_number(latest.get('success'))}, success rate "
            f"{format_number((latest.get('success_rate') or 0) * 100, '%') if latest.get('success_rate') is not None else 'N/A'}."
        )

    if promo.get("available"):
        total = promo.get("total", {})
        lines.append(
            f"- Promotion: tổng cost {format_number(total.get('*Final Cost'))}, "
            f"TPV {format_number(total.get('TPV'))}, cost/TPV "
            f"{format_number((total.get('%Cost/TPV') or 0) * 100, '%') if total.get('%Cost/TPV') is not None else 'N/A'}."
        )

    lines.append("- Hành động đề xuất: ưu tiên kiểm tra nhóm sản phẩm/key source đang kéo giảm TPV và bổ sung KPI target nếu cần chốt báo cáo.")
    return "\n".join(lines)


def build_traffic_answer(summaries: Dict[str, Any]) -> str:
    traffic = summaries.get("traffic", {})
    question = summaries.get("question", "")
    requested_app_id = app_id_from_message(question)
    dimension_key = "zone_id" if wants_zone_breakdown(question) else "source"
    dimension_label = "ZoneID" if dimension_key == "zone_id" else "source"
    traffic_matches = [
        match for match in summaries.get("focus_matches", [])
        if match.get("dataset") == "insurance_traffic"
    ]
    traffic_rows = summaries.get("debug_rows", {}).get("insurance_traffic", [])
    table_info = summaries.get("tables", {}).get("insurance_traffic", {})
    traffic_live_view = summaries.get("data_source", {}).get("tableau", {}).get("views", {}).get("insurance_traffic_csv", {})
    live_filter_verified = bool(traffic_live_view.get("filters")) and traffic_live_view.get("filter_verified") is not False
    if not traffic.get("available"):
        return (
            "Chưa có dữ liệu traffic để chẩn đoán funnel.\n"
            "- Cần CSV từ Insurance Traffic Dashboard có Load/Access, Cashier hoặc Success.\n"
            "- Khi có dữ liệu, agent sẽ tách vấn đề theo Load → Cashier → Success."
        )

    scoped_rows = traffic_rows
    if requested_app_id:
        scoped_rows = rows_for_app_id(traffic_rows, requested_app_id)
        if not scoped_rows and live_filter_verified:
            scoped_rows = traffic_rows

    if requested_app_id and not scoped_rows:
        columns = ", ".join(table_info.get("columns", [])[:10])
        return (
            f"Chưa đủ dữ liệu traffic riêng cho AppID {requested_app_id}.\n"
            "- Dữ liệu traffic hiện tại chưa có dòng nào khớp đúng AppID/SKU được hỏi, nên Bé Hadi không dùng số tổng dashboard để kết luận.\n"
            f"- Các field đang có: {columns}.\n"
            "- Cần Tableau trả CSV sau khi filter AppID/SKU, hoặc export detail `Traffic by Source`/`Traffic by ZoneID` có field AppID/SKU."
        )

    if scoped_rows:
        scoped_traffic = summarize_traffic(scoped_rows)
        if scoped_traffic.get("available"):
            traffic = scoped_traffic
    breakdown = build_breakdown(scoped_rows, [dimension_key, "source", "zone_id", "sku", "app_name"])

    latest = traffic["latest"]
    previous = traffic.get("previous") or {}
    scope = f"AppID/SKU {requested_app_id}" if requested_app_id else "All"
    lines = [f"Chẩn đoán Traffic Funnel - {scope}"]
    lines.append(
        f"- Load ngày {latest.get('time')}: {format_number(latest.get('load'))}, "
        f"thay đổi {pct_text(traffic.get('load_change'))} so với {previous.get('time', 'ngày trước')}."
    )
    if latest.get("cashier") is not None:
        lines.append(f"- Cashier: {format_number(latest.get('cashier'))}.")
    lines.append(
        f"- Success: {format_number(latest.get('success'))}, thay đổi {pct_text(traffic.get('success_change'))}; "
        f"success rate hiện tại {format_number((latest.get('success_rate') or 0) * 100, '%') if latest.get('success_rate') is not None else 'N/A'}."
    )
    if breakdown.get("available"):
        lines.append(f"- Breakdown mặc định theo {dimension_label} ngày/kỳ gần nhất:")
        for item in breakdown.get("items", [])[:12]:
            item_latest = item.get("latest", {})
            lines.append(
                f"  - {item.get('name')}: Load {format_number(item_latest.get('load'))}, "
                f"Cashier {format_number(item_latest.get('cashier'))}, "
                f"Success {format_number(item_latest.get('success'))}, "
                f"success rate {format_number((item_latest.get('success_rate') or 0) * 100, '%') if item_latest.get('success_rate') is not None else 'N/A'}."
            )
    elif requested_app_id and live_filter_verified:
        lines.append(f"- Tableau live đã filter {traffic_live_view.get('filters')} cho AppID/SKU {requested_app_id}; view hiện trả aggregate sau filter nên không có breakdown source/zone.")
    lines.append(
        "- Nhận định: nếu Load tăng nhưng Success không tăng tương ứng, vấn đề nghiêng về conversion/payment; "
        "nếu Load giảm, cần kiểm tra traffic source hoặc campaign source."
    )
    if not breakdown.get("available"):
        lines.append(f"- Hành động đề xuất: cần view/export có field {dimension_label} nếu bà muốn Bé Hadi liệt kê từng {dimension_label}.")
    return "\n".join(lines)


def build_incentive_answer(summaries: Dict[str, Any]) -> str:
    promo = summaries.get("promotion", {})
    question = summaries.get("question", "")
    requested_app_id = app_id_from_message(question)
    promotion_matches = [
        match for match in summaries.get("focus_matches", [])
        if match.get("dataset") == "promotion_summary"
    ]
    if not promo.get("available"):
        if requested_app_id:
            return (
                f"Để tính promotion cost riêng cho AppID {requested_app_id}, Bé Hadi cần thêm dữ liệu promotion đúng scope.\n"
                f"- Bà gửi giúp promotion/campaign code đang chạy cho AppID {requested_app_id}, hoặc export Promotion Summary đã filter theo AppID/code đó.\n"
                "- Field tối thiểu cần có: Campaign Code hoặc Campaign ID, App ID, *Final Cost, TPV, %Cost/TPV, Promo Users/Promo Trans nếu có."
            )
        return (
            "Chưa có dữ liệu Promotion Summary để đánh giá incentive.\n"
            "- Cần các field cost, TPV, cost/TPV, CPU/CAC, promo users và promo transactions.\n"
            "- Nếu có promotion code list theo tháng, hãy paste vào để agent track đúng campaign."
        )

    total = promo.get("total", {})
    lines = ["Đánh giá Incentive/Promotion"]
    if requested_app_id and not promotion_matches:
        lines.append(
            f"- Bé Hadi thấy bà đang hỏi riêng AppID {requested_app_id}, nhưng Promotion Summary hiện chưa có row/code khớp AppID này."
        )
        lines.append(
            f"- Cần bà gửi thêm promotion/campaign code của AppID {requested_app_id}, hoặc để Tableau export có field App ID/Campaign Code sau khi filter."
        )
        lines.append("- Phần dưới chỉ là tổng promotion hiện có trong data, không được xem là cost riêng của AppID này.")
    lines.append(
        f"- Tổng cost: {format_number(total.get('*Final Cost'))}; TPV ghi nhận: {format_number(total.get('TPV'))}; "
        f"cost/TPV: {format_number((total.get('%Cost/TPV') or 0) * 100, '%') if total.get('%Cost/TPV') is not None else 'N/A'}."
    )
    lines.append(
        f"- Promo users: {format_number(total.get('Promo Users'))}; promo transactions: {format_number(total.get('Promo Trans'))}; "
        f"AOV: {format_number(total.get('AOV'))}."
    )
    if promo.get("top_groups"):
        top = promo["top_groups"][0]
        lines.append(
            f"- Nhóm đóng góp TPV cao nhất trong sample: {top['group']} với TPV {format_number(top.get('tpv'))}, "
            f"cost {format_number(top.get('final_cost'))}."
        )
    lines.append(
        "- Nhận định: cost/TPV quanh 50% là cao, nên cần soi lại mục tiêu campaign: acquisition, retention hay subsidy bắt buộc."
    )
    lines.append("- Hành động đề xuất: tối ưu hoặc dừng các campaign có cost/TPV cao nhưng TPV/MPU thấp; giữ campaign nếu chứng minh được uplift tốt.")
    return "\n".join(lines)


def build_product_answer(summaries: Dict[str, Any]) -> str:
    perf = summaries.get("performance", {})
    question = summaries.get("question", "")
    requested_app_id = app_id_from_message(question)
    requested_metrics = requested_metric_keys(question)
    lines = ["Chẩn đoán Performance"]
    breakdown = summaries.get("breakdowns", {}).get("product_performance", {})
    focus_matches = [
        match for match in summaries.get("focus_matches", [])
        if match.get("dataset") in {"insurance_product_performance", "insurance_performance"}
    ]
    source_info = summaries.get("data_source", {})

    if focus_matches:
        focus_values = set()
        for match in focus_matches:
            for value in match.get("matched_fields", {}).values():
                normalized = normalize_text(value)
                if normalized:
                    focus_values.add(normalized)
                digits = digits_only(value)
                if digits:
                    focus_values.add(digits)

        focused_items = []
        if breakdown.get("available"):
            for item in breakdown.get("items", []):
                name = item.get("name", "")
                normalized_name = normalize_text(name)
                digit_name = digits_only(name)
                if normalized_name in focus_values or digit_name in focus_values:
                    focused_items.append(item)

        if focused_items:
            lines.append("- Kết quả đúng AppID/SKU được hỏi:")
            focused_notes = []
            focused_actions = []
            for item in focused_items[:3]:
                latest = item.get("latest", {})
                previous = item.get("previous", {})
                tpv_change = item.get("tpv_change")
                mpu_change = item.get("mpu_change")
                if mpu_change is not None and mpu_change <= -0.15:
                    focused_health = "cần chú ý vì MPU giảm mạnh"
                elif tpv_change is not None and tpv_change <= -0.15:
                    focused_health = "báo động vì TPV giảm mạnh"
                elif tpv_change is not None and tpv_change >= 0 and (mpu_change is None or mpu_change >= 0):
                    focused_health = "tốt"
                else:
                    focused_health = "cần theo dõi thêm"
                metric_parts = metric_parts_for_answer(latest, requested_metrics)
                lines.append(
                    f"  - {item.get('name')}: kỳ {item.get('latest_period')} đạt "
                    f"{', '.join(metric_parts)}."
                )
                change_parts = metric_change_parts(item, requested_metrics)
                lines.append(
                    f"    So với {item.get('previous_period') or 'kỳ trước'}: {', '.join(change_parts)}, "
                    f"trạng thái {focused_health}."
                )
                if previous:
                    previous_parts = metric_parts_for_answer(previous, requested_metrics)
                    lines.append(
                        f"    Baseline trước đó: {', '.join(previous_parts)}."
                    )
                if tpv_change is not None and tpv_change >= 0 and (mpu_change is None or mpu_change >= 0):
                    focused_notes.append(
                        f"{item.get('name')} đang tích cực: TPV và MPU đều tăng hoặc giữ được nền."
                    )
                    focused_actions.append(
                        "tiếp tục theo dõi source/campaign để biết tăng trưởng đến từ organic hay incentive."
                    )
                elif mpu_change is not None and mpu_change < 0 <= (tpv_change or 0):
                    focused_notes.append(
                        f"{item.get('name')} có TPV tăng nhưng MPU giảm, tăng trưởng có thể đến từ AOV/mix giao dịch."
                    )
                    focused_actions.append(
                        "kiểm tra conversion funnel và nhóm user/source bị hụt để tránh mất nền người mua."
                    )
                elif tpv_change is not None and tpv_change < 0:
                    focused_notes.append(
                        f"{item.get('name')} đang giảm TPV, cần soi thêm traffic source, campaign và conversion."
                    )
                    focused_actions.append(
                        "ưu tiên đối chiếu traffic, promotion cost và thay đổi SKU/campaign trong kỳ."
                    )
                else:
                    focused_notes.append(
                        f"{item.get('name')} chưa đủ baseline để kết luận chắc chắn."
                    )
                    focused_actions.append(
                        "bổ sung thêm kỳ trước hoặc daily MTD để đánh giá đúng xu hướng."
                    )
            if focused_notes:
                lines.append(f"- Nhận định: {focused_notes[0]}")
            if focused_actions:
                lines.append(f"- Hành động đề xuất: {focused_actions[0]}")
            return "\n".join(lines)

        lines.append("- Agent tìm thấy dữ liệu khớp với sản phẩm/AppID/SKU trong câu hỏi:")
        for match in focus_matches[:4]:
            metrics = match.get("metrics", {})
            lines.append(
                f"  - {match.get('matched_fields')} | period {match.get('period')} | "
                f"{', '.join(metric_parts_for_answer(metrics, requested_metrics))}."
            )
        lines.append("- Chưa đủ dữ liệu nhiều kỳ để tính tăng/giảm cho đúng AppID/SKU này.")
        return "\n".join(lines)

    if breakdown.get("available"):
        if requested_app_id:
            tableau = source_info.get("tableau", {})
            performance_view = tableau.get("views", {}).get("insurance_performance_csv", {})
            product_view = tableau.get("views", {}).get("insurance_product_performance_csv", {})
            verified_view = product_view if product_view.get("filters") else performance_view
            if verified_view.get("filters") and verified_view.get("filter_verified") is not False:
                lines.append(
                    f"- Tableau live đã filter {verified_view.get('filters')} cho AppID/SKU {requested_app_id}."
                )
                if breakdown.get("total_groups") == 1:
                    item = (breakdown.get("items") or [{}])[0]
                    latest = item.get("latest", {})
                    previous = item.get("previous", {})
                    metric_parts = metric_parts_for_answer(latest, requested_metrics)
                    change_parts = metric_change_parts(item, requested_metrics)
                    lines.append(
                        f"- Kết quả sau filter: {item.get('latest_period')} {', '.join(metric_parts)}."
                    )
                    lines.append(
                        f"- So với {item.get('previous_period') or 'kỳ trước'}: {', '.join(change_parts)}; "
                        f"baseline {', '.join(metric_parts_for_answer(previous, requested_metrics))}."
                    )
                    lines.append("- View hiện trả aggregate sau filter nên chưa có breakdown nhỏ hơn.")
                    return "\n".join(lines)
            else:
                columns = ", ".join(summaries.get("tables", {}).get("insurance_product_performance", {}).get("columns", [])[:10])
                return "\n".join(
                    [
                        f"Chưa đủ dữ liệu performance riêng cho AppID/SKU {requested_app_id}.",
                        "- Bé Hadi không dùng số All để kết luận cho một AppID cụ thể.",
                        f"- Field đang có: {columns}.",
                        "- Cần Tableau live filter đúng AppID/SKU hoặc export Break View theo AppID/SKU.",
                    ]
                )
        lines.append(
            f"- Breakdown theo {breakdown.get('dimension_field')} có {breakdown.get('total_groups')} nhóm. Top nhóm theo TPV/MPU:"
        )
        for item in breakdown.get("items", [])[:8]:
            latest = item.get("latest", {})
            lines.append(
                f"  - {item.get('name')}: TPV {format_number(latest.get('tpv'))}, "
                f"MPU {format_number(latest.get('mpu'))}, "
                f"TPV change {pct_text(item.get('tpv_change'))}, trạng thái {item.get('health')}."
            )
        lines.append("- Nếu muốn soi một AppID cụ thể, hỏi trực tiếp mã AppID/SKU để agent lọc đúng row liên quan.")
        return "\n".join(lines)

    if perf.get("available"):
        latest = perf["latest"]
        previous = perf.get("previous") or {}
        tableau = source_info.get("tableau", {})
        performance_view = tableau.get("views", {}).get("insurance_performance_csv", {})
        performance_filters = performance_view.get("filters", {})
        if requested_app_id and performance_filters and performance_view.get("filter_verified") is not False:
            lines.append(
                f"- Tableau live đã gọi Insurance Performance với filter {performance_filters} cho AppID/SKU {requested_app_id}."
            )
            latest_metrics = {"tpv": latest.get("tpv"), "mpu": latest.get("users"), "aov": latest.get("aov")}
            previous_metrics = {"tpv": previous.get("tpv"), "mpu": previous.get("users"), "aov": previous.get("aov")}
            changes = []
            if "tpv" in requested_metrics:
                changes.append(f"TPV change {pct_text(perf.get('tpv_change'))}")
            if "mpu" in requested_metrics:
                changes.append(f"MPU change {pct_text(perf.get('user_change'))}")
            lines.append(
                f"- Kết quả view sau filter: {latest.get('period')} {', '.join(metric_parts_for_answer(latest_metrics, requested_metrics))}; "
                f"so với {previous.get('period', 'kỳ trước')} {', '.join(changes)}, "
                f"trạng thái {perf.get('health')}."
            )
            lines.append(f"- Baseline: {', '.join(metric_parts_for_answer(previous_metrics, requested_metrics))}.")
            lines.append(
                "- Nếu số này vẫn ra tổng All, cần đổi `TABLEAU_PERFORMANCE_APPID_FILTER_FIELDS` sang đúng tên filter trên Atlas hoặc cung cấp exact SKU value."
            )
            return "\n".join(lines)
        if requested_app_id and performance_filters and performance_view.get("filter_verified") is False:
            lines.append(
                f"- Bé Hadi đã thử filter Tableau cho AppID/SKU {requested_app_id}, nhưng CSV trả về giống bản tổng All nên chưa xem là dữ liệu riêng AppID."
            )
            lines.append(
                "- Cần đổi đúng filter caption trong `TABLEAU_PERFORMANCE_APPID_FILTER_FIELDS` hoặc gửi exact SKU value đang hiển thị trong dropdown Atlas."
            )
            return "\n".join(lines)
        lines.append(
            f"- Dữ liệu hiện tại đang ở mức aggregate: {latest.get('period')} TPV {format_number(latest.get('tpv'))}, "
            f"giảm/tăng {pct_text(perf.get('tpv_change'))} so với {previous.get('period', 'kỳ trước')}."
        )
        lines.append(
            "- Với dữ liệu aggregate này, agent xác định được business đang giảm, nhưng chưa đủ để kết luận sản phẩm cụ thể nào kéo giảm."
        )
    else:
        lines.append("- Chưa có CSV từ Insurance Performance để phân tích sản phẩm.")
    if source_info.get("tableau_live_requested"):
        tableau = source_info.get("tableau", {})
        if tableau.get("used"):
            lines.append("- Nguồn dữ liệu hiện tại: Tableau live export. View đang trả aggregate nếu không thấy AppID/SKU.")
        if tableau.get("errors"):
            tableau_errors = tableau.get("errors", [])[:2]
            lines.append(f"- Tableau live có lỗi: {'; '.join(tableau_errors)}.")
            if any("403" in str(error) or "Forbidden" in str(error) for error in tableau_errors):
                lines.append(
                    "- Lỗi này thường do AgentBase public runtime bị Atlas/nginx chặn khi gọi Tableau trực tiếp. "
                    "Cần chạy runtime trong VPC có quyền vào Atlas hoặc cấu hình `TABLEAU_PROXY_URL` trỏ tới bridge HTTPS từ máy có VPN."
                )
                return "\n".join(lines)
    missing_reason = breakdown.get("reason") if breakdown else None
    lines.append(
        "- Cần export thêm breakdown theo Product Name/App Name/SKU/App ID. "
        f"{missing_reason or 'Hiện payload chưa đủ dimension để trả lời theo từng sản phẩm.'}"
    )
    lines.append("- Hành động đề xuất: lọc dashboard theo Product Name hoặc Break View = product/SKU rồi paste CSV vào payload.")
    return "\n".join(lines)


KEY_REPORT_PRODUCTS = [
    {"code": "4491", "owner": "Hân", "name": "Bảo hiểm Sống An Vui ChubbGI"},
    {"code": "3507", "owner": "Hân", "name": "Bảo hiểm Savemoney Ô tô TNDS"},
    {"code": "4326", "owner": "Hân", "name": "Bảo hiểm Trách nhiệm dân sự bắt buộc"},
    {"code": "3394", "owner": "Trân", "name": "Bảo hiểm màn hình điện thoại"},
    {"code": "3682", "owner": "Trân", "name": "Bảo hiểm an ninh mạng VBI"},
    {"code": "3274", "owner": "Trân", "name": "Bảo hiểm du lịch nội địa"},
]


def find_breakdown_item_by_code(breakdown: Dict[str, Any], code: str) -> Optional[Dict[str, Any]]:
    for item in breakdown.get("items", []):
        if code in digits_only(item.get("name", "")):
            return item
        raw_latest = item.get("raw_latest", {})
        if row_matches_app_id(raw_latest, code):
            return item
    return None


def build_daily_email_report(summaries: Dict[str, Any]) -> str:
    performance = summaries.get("performance", {})
    traffic = summaries.get("traffic", {})
    product_breakdown = summaries.get("breakdowns", {}).get("product_performance", {})
    source_breakdown = summaries.get("breakdowns", {}).get("traffic_source", {})
    latest_period = "T-1"
    if performance.get("latest", {}).get("cutoff"):
        latest_period = performance["latest"]["cutoff"]
    elif performance.get("latest", {}).get("period"):
        latest_period = performance["latest"]["period"]

    lines = [
        f"[Bé Hadi] Daily Insurance Pulse - {latest_period}",
        "",
        "1. Performance key AppID",
    ]
    for owner in ["Hân", "Trân"]:
        lines.append(f"{owner}:")
        owner_products = [product for product in KEY_REPORT_PRODUCTS if product["owner"] == owner]
        for product in owner_products:
            item = find_breakdown_item_by_code(product_breakdown, product["code"]) if product_breakdown.get("available") else None
            if item:
                latest = item.get("latest", {})
                lines.append(
                    f"- {product['code']} - {product['name']}: TPV {format_number(latest.get('tpv'))}, "
                    f"MPU {format_number(latest.get('mpu'))}, TPV change {pct_text(item.get('tpv_change'))}, "
                    f"MPU change {pct_text(item.get('mpu_change'))}, trạng thái {item.get('health')}."
                )
            else:
                lines.append(f"- {product['code']} - {product['name']}: chưa có row AppID/SKU trong export hiện tại.")

    lines.extend(["", "2. Traffic"])
    if traffic.get("available"):
        latest = traffic.get("latest", {})
        lines.append(
            f"- All: Load {format_number(latest.get('load'))}, Cashier {format_number(latest.get('cashier'))}, "
            f"Success {format_number(latest.get('success'))}, Success rate "
            f"{format_number((latest.get('success_rate') or 0) * 100, '%') if latest.get('success_rate') is not None else 'N/A'}."
        )
    else:
        lines.append("- Chưa có dữ liệu traffic.")

    if source_breakdown.get("available"):
        top_sources = sorted(
            source_breakdown.get("items", []),
            key=lambda item: item.get("latest", {}).get("success") or 0,
            reverse=True,
        )[:3]
        lines.append("- Top 3 source theo Success:")
        for item in top_sources:
            latest = item.get("latest", {})
            lines.append(
                f"  - {item.get('name')}: Load {format_number(latest.get('load'))}, "
                f"Cashier {format_number(latest.get('cashier'))}, Success {format_number(latest.get('success'))}."
            )
    else:
        lines.append("- Chưa có breakdown source để xếp top 3 source theo Success.")

    lines.extend([
        "",
        "3. Ghi chú",
        "- Báo cáo dùng số T-1/gần nhất mà Tableau trả về.",
        "- Nếu một AppID thiếu row, cần kiểm tra filter AppID/SKU hoặc export Break View từ dashboard Performance/Traffic.",
    ])
    return "\n".join(lines)


def build_report_answer(summaries: Dict[str, Any]) -> str:
    lines = ["Draft Monthly Report - Insurance"]
    lines.append("1. Executive Summary")
    perf = summaries.get("performance", {})
    if perf.get("available"):
        latest = perf["latest"]
        lines.append(
            f"- Tính đến {latest.get('cutoff') or latest.get('period')}, TPV đạt {format_number(latest.get('tpv'))}, "
            f"users {format_number(latest.get('users'))}, trạng thái {perf.get('health')}."
        )
    else:
        lines.append("- Chưa có dữ liệu performance để điền executive summary.")
    lines.append("2. What Went Well")
    lines.append("- Traffic ngày gần nhất có tín hiệu cải thiện nếu success/users tăng so với ngày trước.")
    lines.append("3. What Didn't Go Well")
    lines.append("- TPV MoM đang cần chú ý nếu mức giảm vượt 15% theo rule business.")
    lines.append("4. Recommended Actions")
    lines.append("- Bổ sung KPI target, incentive objective và promotion code list để hoàn thiện bản report cuối cùng.")
    return "\n".join(lines)


def build_deterministic_answer(message: str, summaries: Dict[str, Any], mode: str = "") -> str:
    intent = classify_intent(message, mode)
    if intent == "traffic":
        return build_traffic_answer(summaries)
    if intent == "incentive":
        return build_incentive_answer(summaries)
    if intent in {"product", "performance"}:
        return build_product_answer(summaries)
    if intent == "report":
        return build_report_answer(summaries)
    return build_daily_answer(summaries)


def call_llm_if_configured(message: str, summaries: Dict[str, Any], mode: str = "") -> Dict[str, Any]:
    api_key = env_value("LLM_API_KEY", "AI_PLATFORM_API_KEY")
    base_url = normalize_llm_base_url(env_value("LLM_BASE_URL", default="https://maas-llm-aiplatform-hcm.api.vngcloud.vn/v1"))
    model = normalize_llm_model(env_value("LLM_MODEL", default="qwen/qwen3-5-27b"))
    timeout_seconds = parse_number(os.getenv("LLM_TIMEOUT_SECONDS")) or 90
    if not (api_key and base_url and model and OpenAI):
        return {"response": None, "error": None}

    try:
        client = OpenAI(api_key=api_key, base_url=base_url, timeout=timeout_seconds, max_retries=1)
        prompt = f"""
You are Bé Hadi Bảo hiểm, an AI Business Analyst for insurance products.
Always answer in Vietnamese with full accents. Use a concise, business-friendly tone.

Knowledge base:
{KNOWLEDGE_BASE.as_prompt_context(max_chars=12000)}

User question:
{message}

Selected UI mode:
{mode or classify_intent(message)}

Computed data summaries:
{json.dumps(summaries, ensure_ascii=False, indent=2)}

Answer rules:
- Use the computed summaries as the source of truth. Do not invent numbers.
- Current product scope focuses on Performance and Traffic dashboards. Do not lead with Incentive/Promotion unless the user explicitly asks.
- AppID and SKU can overlap. If the user provides a numeric AppID, treat it as an AppID/SKU filter candidate. If the user writes a SKU name, match textually.
- Change/MoM formula is always (current - previous) / previous * 100. Current is X, previous is Y.
- If the user asks MTD/by product/by AppID/by SKU, prioritize Tableau filtered rows, `breakdowns.product_performance`, and `focus_matches`.
- If a specific AppID/SKU/product appears in `focus_matches`, answer specifically for that item before giving broader context.
- If `focus_matches` is not empty, do not list unrelated AppIDs/products unless the user explicitly asks for top products, all products, comparison, or leaderboard.
- For traffic, default to breakdown by source. Only use breakdown by zone when the user asks for zone/ZoneID.
- For traffic by AppID/SKU, do not use All/overview traffic unless Tableau metadata confirms the view was filtered to that AppID/SKU.
- If the needed dimension is missing from the uploaded/exported CSV, say exactly which field is missing and what Tableau export/filter is needed.
- If the user asks promotion cost for a specific AppID but no campaign/promotion code or AppID-level Promotion Summary rows are present, ask for the campaign code or a Tableau export filtered to that AppID. Do not treat total `All` promotion cost as that AppID's cost.
- For product or AppID breakdowns, include a compact table-style list with product/AppID/SKU, latest TPV/MPU, change vs previous period, and status.
- Apply business rules from the knowledge base: MoM drop >5% is not good, drop >=15% is alerting, MTD below same-day previous month is not good.
- Ignore MWEB/Others by default unless the user explicitly asks about it.
- If uploaded_files are available, use their extracted text as additional business/data context. If the file contains rows or report content that directly answer the user question, summarize that content clearly and state that the answer uses the uploaded file.

Suggested answer shape:
1. Câu trả lời ngắn
2. Số liệu chính
3. Bất thường/rủi ro
4. Hành động đề xuất
"""
        response = client.chat.completions.create(
            model=model,
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are Bé Hadi Bảo hiểm, a careful Vietnamese AI Business Analyst for insurance. "
                        "Be specific to the user's context and use only provided data."
                    ),
                },
                {"role": "user", "content": prompt},
            ],
            temperature=0.35,
            max_tokens=1200,
        )
        return {"response": response.choices[0].message.content, "error": None}
    except Exception as exc:
        return {
            "response": None,
            "error": {
                "type": exc.__class__.__name__,
                "message": str(exc)[:500],
            },
        }


def handle_payload(payload: dict, context: RequestContext) -> dict:
    message = payload.get("message") or payload.get("question") or "Analyze insurance performance."
    mode = payload.get("mode", "")
    data = payload.get("data") or {}
    fallback_data = data
    if "use_tableau_live" in payload:
        use_tableau_live = bool(payload.get("use_tableau_live"))
    else:
        use_tableau_live = bool_env("TABLEAU_AUTO_REFRESH", default=False)
    tableau_metadata = {"enabled": False, "used": False, "errors": [], "views": {}}

    if payload.get("action") == "knowledge_status" or message.lower() in {
        "knowledge_status",
        "kb_status",
        "status knowledge base",
    }:
        return {
            "status": "success",
            "agent": "be-hadi-bao-hiem",
            "knowledge_base": KNOWLEDGE_BASE.status(),
            "timestamp": datetime.now().isoformat(),
            "session_id": context.session_id,
        }

    if payload.get("action") == "daily_email_report":
        payload = dict(payload)
        payload["message"] = payload.get("message") or "Tạo daily report T-1 cho Insurance Performance và Traffic."
        payload["mode"] = "performance"

    if payload.get("action") == "tableau_status" or message.lower() in {
        "tableau_status",
        "status tableau",
        "tableau live status",
    }:
        live_result = fetch_tableau_live_data(message)
        metadata = live_result["metadata"]
        return {
            "status": "success" if metadata.get("used") else "error",
            "agent": "be-hadi-bao-hiem",
            "tableau_live_used": metadata.get("used", False),
            "tableau": metadata,
            "tables": {
                key: {"row_count": len(parse_csv_text(value))}
                for key, value in live_result.get("data", {}).items()
            },
            "timestamp": datetime.now().isoformat(),
            "session_id": context.session_id,
        }

    if use_tableau_live:
        live_result = fetch_tableau_live_data(message)
        tableau_metadata = live_result["metadata"]
        if live_result["data"]:
            data = live_result["data"]
        elif bool_env("ALLOW_DEMO_DATA_WHEN_TABLEAU_FAILS", default=False):
            data = fallback_data
        else:
            data = {}

    uploaded_files = summarize_uploaded_files(payload.get("uploaded_files"))
    performance_rows = parse_csv_text(data.get("insurance_performance_csv", ""))
    product_performance_rows = parse_csv_text(data.get("insurance_product_performance_csv", ""))
    traffic_detail_rows = parse_csv_text(data.get("insurance_traffic_detail_csv", ""))
    traffic_raw_rows = traffic_detail_rows or parse_csv_text(data.get("insurance_traffic_csv", ""))
    traffic_rows = normalize_traffic_crosstab_rows(traffic_raw_rows) or traffic_raw_rows
    promotion_rows = parse_csv_text(data.get("promotion_summary_csv", ""))
    product_rows = product_performance_rows or performance_rows
    datasets = {
        "insurance_performance": performance_rows,
        "insurance_product_performance": product_rows,
        "insurance_traffic": traffic_rows,
        "promotion_summary": promotion_rows,
    }

    summaries = {
        "question": message,
        "performance": summarize_performance(performance_rows),
        "traffic": summarize_traffic(traffic_rows),
        "promotion": summarize_promotion(promotion_rows),
        "uploaded_files": uploaded_files,
        "data_source": {
            "tableau_live_requested": use_tableau_live,
            "tableau_live_used": tableau_metadata.get("used", False),
            "tableau": tableau_metadata,
        },
        "tables": {
            "insurance_performance": table_profile(performance_rows),
            "insurance_product_performance": table_profile(product_rows),
            "insurance_traffic": table_profile(traffic_rows),
            "insurance_traffic_detail": table_profile(traffic_detail_rows),
            "promotion_summary": table_profile(promotion_rows),
        },
        "debug_rows": {
            "insurance_traffic": traffic_rows[:300],
        },
        "breakdowns": {
            "product_performance": build_breakdown(
                product_rows,
                ["app_id", "sku", "product_name", "app_name", "break_view"],
            ),
            "traffic_source": build_breakdown(
                traffic_rows,
                ["source", "sku", "app_id", "app_name", "product_name"],
            ),
            "promotion_campaign": build_breakdown(
                promotion_rows,
                ["campaign", "app_id", "source", "product_name"],
            ),
        },
        "focus_matches": find_focus_matches(message, datasets),
    }

    if payload.get("action") == "daily_email_report":
        response = build_daily_email_report(summaries)
        return {
            "status": "success",
            "agent": "be-hadi-bao-hiem",
            "mode": "daily_email_report",
            "response": response,
            "email": {
                "subject": f"[Bé Hadi] Daily Insurance Pulse - {datetime.now().strftime('%Y-%m-%d')}",
                "recipients": [
                    "hanlgb@vng.com.vn",
                    "mynt5@vng.com.vn",
                    "tramntq@vng.com.vn",
                    "tranvhd@vng.com.vn",
                ],
            },
            "llm": {
                "configured": bool(env_value("LLM_API_KEY", "AI_PLATFORM_API_KEY")),
                "used": False,
                "error": None,
            },
            "data_source": summaries["data_source"],
            "knowledge_base": KNOWLEDGE_BASE.status(),
            "summaries": summaries,
            "timestamp": datetime.now().isoformat(),
            "session_id": context.session_id,
        }

    resolved_mode = classify_intent(message, mode)
    llm_result = call_llm_if_configured(message, summaries, resolved_mode)
    response = llm_result.get("response") or build_deterministic_answer(message, summaries, resolved_mode)

    return {
        "status": "success",
        "agent": "be-hadi-bao-hiem",
        "mode": resolved_mode,
        "response": response,
        "llm": {
            "configured": bool(env_value("LLM_API_KEY", "AI_PLATFORM_API_KEY")),
            "used": bool(llm_result.get("response")),
            "error": llm_result.get("error"),
        },
        "data_source": summaries["data_source"],
        "knowledge_base": KNOWLEDGE_BASE.status(),
        "summaries": summaries,
        "timestamp": datetime.now().isoformat(),
        "session_id": context.session_id,
    }


if app:
    @app.entrypoint
    def handler(payload: dict, context: RequestContext) -> dict:
        return handle_payload(payload, context)


    @app.ping
    def health_check() -> PingStatus:
        return PingStatus.HEALTHY
else:
    from fastapi import FastAPI, Request
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.responses import FileResponse

    app = FastAPI(title="Bé Hadi Bảo hiểm Local Preview")
    app.add_middleware(
        CORSMiddleware,
        allow_origins=["*"],
        allow_credentials=False,
        allow_methods=["*"],
        allow_headers=["*"],
    )

    @app.get("/health")
    async def health_check() -> Dict[str, str]:
        return {"status": "healthy", "mode": "fastapi-cors-runtime"}

    @app.get("/")
    async def preview() -> FileResponse:
        return FileResponse(ROOT / "preview.html")

    @app.get("/data_exports/tableau-live-payload.json")
    async def tableau_live_payload() -> FileResponse:
        return FileResponse(ROOT / "data_exports" / "tableau-live-payload.json")

    @app.options("/{path:path}")
    async def options_preflight(path: str) -> Dict[str, str]:
        return {"status": "ok"}

    @app.post("/invocations")
    async def invoke(request: Request) -> Dict[str, Any]:
        payload = await request.json()
        context = SimpleNamespace(
            session_id=request.headers.get("X-GreenNode-AgentBase-Session-Id", "local-session"),
            user_id=request.headers.get("X-GreenNode-AgentBase-User-Id", "local-user"),
            request_id=request.headers.get("X-GreenNode-AgentBase-Request-Id", "local-request"),
            request_headers=dict(request.headers),
        )
        response = handle_payload(payload, context)
        response["runtime_mode"] = "fastapi-cors-runtime"
        if AGENTBASE_IMPORT_ERROR:
            response["agentbase_import_error"] = str(AGENTBASE_IMPORT_ERROR)
        return response


if __name__ == "__main__":
    if USE_GREENNODE_AGENTBASE_APP and GreenNodeAgentBaseApp:
        app.run(port=8080, host="0.0.0.0")
    else:
        import uvicorn

        uvicorn.run(app, host="0.0.0.0", port=8080)
